"""
Chat API - Step 6: LangChain + Natural Language Processing
자연어 → QueryPlan → Core API → RenderSpec

Text-to-SQL 모드 추가:
SQL_ENABLE_TEXT_TO_SQL=true 설정 시 AI가 직접 SQL을 생성하여 읽기 전용 DB에서 실행
"""

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import Optional, Dict, Any, List, Generator, Tuple
from enum import Enum
import httpx
import logging
import os
import re
import uuid
import csv
import io
import math
from datetime import datetime

from app.constants.render_keywords import (
    CHART_KEYWORDS,
    TABLE_KEYWORDS,
    TIME_FIELD_KEYWORDS,
)


# ============================================
# 참조 표현 감지 (연속 대화 WHERE 조건 병합용)
# ============================================

class ReferenceType(str, Enum):
    """참조 표현 유형"""
    LATEST = "latest"      # 직전 결과 참조 ("이중에", "방금")
    SPECIFIC = "specific"  # 특정 결과 참조 ("30건에서", "첫 번째 결과")
    PARTIAL = "partial"    # 부분 결과 참조 ("상위 10개", "처음 5건")
    NONE = "none"          # 참조 표현 없음


# 확장된 참조 표현 패턴 (30개+)
# Phase 3: 암시적 필터 패턴 추가 (4단계+ 체이닝 지원)
REFERENCE_PATTERNS = {
    ReferenceType.LATEST: [
        # 한글 - 표준 표현
        r'이\s*중에?서?',        # "이중에", "이 중에서", "이중에서"
        r'여기서',               # "여기서"
        r'그\s*중에?서?',        # "그중에", "그 중에서"
        r'직전\s*(결과|데이터)?', # "직전", "직전 결과"
        r'방금\s*(결과|데이터)?', # "방금", "방금 결과"
        r'위\s*결과',            # "위 결과", "위결과"
        r'앞\s*(서|에서)',        # "앞서", "앞에서"
        r'해당\s*데이터',         # "해당 데이터"
        r'이\s*결과에?서?',       # "이 결과", "이 결과에서"
        r'저\s*중에?서?',         # "저중에", "저 중에서"
        r'거기서',               # "거기서"
        # 한글 - 구어체/줄임말
        r'아까\s*(그\s*)?(거|것|데이터)?', # "아까 그거", "아까 거"
        r'그거에?서?',           # "그거에서", "그거서"
        r'이거에?서?',           # "이거에서", "이거서"
        r'조회\s*한\s*거에?서?',  # "조회한 거에서"
        r'나온\s*(거|것)에?서?',  # "나온 거에서"
        r'보여준\s*(거|것)에?서?', # "보여준 거에서"
        r'받은\s*(거|것|데이터)에?서?', # "받은 거에서"
        r'화면\s*(에\s*)?(있는|보이는)', # "화면에 있는"
        r'지금\s*(보이는|있는)',  # "지금 보이는"
        # 한글 - 상황별 표현
        r'조회된\s*결과',         # "조회된 결과"
        r'이전\s*결과에?서?',     # "이전 결과에서"
        r'검색\s*결과에?서?',     # "검색 결과에서"
        r'목록에?서?',           # "목록에서"
        r'테이블에?서?',         # "테이블에서"
        # 영어/영한 혼용
        r'from\s*(this|these|here)', # "from this", "from these"
        r'among\s*(this|these)',     # "among these"
        r'in\s*this\s*(result|data|list)', # "in this result"
        r'out\s*of\s*(this|these)',  # "out of these"
        r'(this|these)\s*중에?서?',  # "these 중에서"

        # Phase 3: 암시적 필터 패턴 (문장 끝 "~만" 표현)
        # 4단계+ 체이닝에서 "금액 10만원 이상만" 같은 표현 감지
        r'.{2,}(것|건|거|데이터)만\s*$',  # "~것만", "~건만", "~거만" 으로 끝남
        r'.{2,}(인|한|된|는)\s*것만\s*$', # "~인 것만", "~한 것만" 으로 끝남
        r'(금액|amount).{0,15}(이상|이하|초과|미만).{0,5}만', # "금액 X 이상만"
        r'(상태|status).{0,10}(인|가|만)',  # "상태가 X인 것만"
        r'(결제|method).{0,10}(인|가|만)',  # "결제수단이 X인 것만"
        r'(가맹점|merchant).{0,10}(만|것만)', # "가맹점 X만"

        # Phase 3: 암시적 필터 - 비교/범위 표현
        r'(만원|원)\s*(이상|이하|초과|미만)',  # "10만원 이상"
        r'\d+\s*(이상|이하|초과|미만)\s*만?$', # "100 이상만"
        r'(크|작|높|낮|많|적)(은|고)\s*(것|건|거)만',  # "큰 것만", "작은 건만"

        # Phase 3: 필터 추가 표현
        r'추가로\s*.{0,10}(필터|조건)',  # "추가로 필터"
        r'(조건|필터)\s*추가',            # "조건 추가"
        r'더\s*좁혀',                    # "더 좁혀서"
        r'범위\s*좁혀',                  # "범위 좁혀서"
    ],
    ReferenceType.SPECIFIC: [
        # 특정 결과 지정 (숫자 앞에 맥락 필요)
        r'아까\s*\d+건',         # "아까 30건"
        r'(그|저)\s*\d+건에?서?', # "그 30건에서"
        r'(첫|두|세)\s*번째\s*(결과|데이터)', # "첫 번째 결과"
        r'(처음|마지막)\s*(결과|데이터)', # "처음 결과", "마지막 결과"
        r'(이전|앞의?)\s*조회',   # "이전 조회"
        r'결과\s*\d+건에?서?',   # "결과 30건에서"
    ],
    ReferenceType.PARTIAL: [
        # 부분 결과 참조
        r'상위\s*\d+',           # "상위 10개"
        r'하위\s*\d+',           # "하위 5개"
        r'처음\s*\d+건?',        # "처음 5건"
        r'위\s*\d+건?',          # "위 10건"
        r'top\s*\d+',            # "top 10"
        r'first\s*\d+',          # "first 5"
    ]
}

# 새 쿼리 패턴 (이전 조건 무시)
NEW_QUERY_PATTERNS = [
    r'새로\s*.{0,10}조회',   # "새로 조회", "새로 환불 내역 조회"
    r'다시\s*.{0,10}조회',   # "다시 조회", "다시 결제 조회"
    r'처음부터',             # "처음부터"
    r'새\s*쿼리',            # "새 쿼리"
    r'전체\s*다시',          # "전체 다시"
    r'전체\s*조회',          # "전체 조회"
    r'새로\s*검색',          # "새로 검색"
    r'다른\s*(데이터|것|거)', # "다른 데이터"
    r'별도로',               # "별도로"
    r'fresh\s*query',        # "fresh query"
    r'new\s*search',         # "new search"
]

# 집계 키워드 패턴 (이전 결과 참조로 처리)
# 이전 대화에서 조회한 결과에 대해 집계하는 것으로 간주
AGGREGATION_KEYWORDS = [
    # 한글 집계 표현
    r'합산',                 # "합산해줘"
    r'합계',                 # "합계 보여줘"
    r'총\s*(금액|결제|매출|건수|수량)', # "총 금액", "총 결제", "총 매출"
    r'전체\s*(금액|결제|매출)',  # "전체 금액" (단, "전체 조회"는 제외)
    r'더해',                 # "더해줘"
    r'sum',                  # "sum 구해줘"
    r'평균',                 # "평균 구해줘"
    r'평균\s*(금액|결제|매출)',  # "평균 금액"
    r'avg',                  # "avg 구해줘"
    r'개수',                 # "개수 세줘"
    r'몇\s*건',              # "몇 건이야"
    r'카운트',               # "카운트 해줘"
    r'count',                # "count 해줘"
    r'최대\s*(금액|값)',      # "최대 금액"
    r'최소\s*(금액|값)',      # "최소 금액"
    r'max',                  # "max 구해줘"
    r'min',                  # "min 구해줘"
]


def detect_reference_expression(message: str) -> Tuple[bool, str]:
    """
    사용자 메시지에서 참조 표현 감지

    참조 표현이 있으면 이전 WHERE 조건을 유지해야 함을 의미

    Args:
        message: 사용자 메시지

    Returns:
        (is_refinement, ref_type) 튜플
        - is_refinement: True면 이전 조건 유지 필요
        - ref_type: 'filter' (필터 추가), 'aggregation' (집계 요청), 'new' (새 쿼리), 'none' (해당없음)
    """
    # 새 쿼리 패턴 먼저 체크 (우선순위 높음)
    for pattern in NEW_QUERY_PATTERNS:
        if re.search(pattern, message, re.IGNORECASE):
            return (False, 'new')

    # 참조 패턴 체크 (유형별)
    for ref_type, patterns in REFERENCE_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, message, re.IGNORECASE):
                return (True, 'filter')

    # 집계 키워드 체크 (이전 결과에 대한 집계로 처리)
    # 집계 요청은 이전 대화에서 조회한 결과에 대해 수행하는 것으로 간주
    for pattern in AGGREGATION_KEYWORDS:
        if re.search(pattern, message, re.IGNORECASE):
            return (True, 'aggregation')

    return (False, 'none')


def detect_reference_type(message: str) -> ReferenceType:
    """
    참조 표현의 세부 유형 감지

    Args:
        message: 사용자 메시지

    Returns:
        ReferenceType: 참조 유형 (LATEST, SPECIFIC, PARTIAL, NONE)
    """
    # 우선순위: SPECIFIC > PARTIAL > LATEST
    for ref_type in [ReferenceType.SPECIFIC, ReferenceType.PARTIAL, ReferenceType.LATEST]:
        if ref_type in REFERENCE_PATTERNS:
            for pattern in REFERENCE_PATTERNS[ref_type]:
                if re.search(pattern, message, re.IGNORECASE):
                    return ref_type

    return ReferenceType.NONE

# Text-to-SQL 모드 플래그
ENABLE_TEXT_TO_SQL = os.getenv("SQL_ENABLE_TEXT_TO_SQL", "false").lower() == "true"


def to_camel(string: str) -> str:
    """snake_case를 camelCase로 변환"""
    components = string.split('_')
    return components[0] + ''.join(x.title() for x in components[1:])


def summarize_query_plan(query_plan: Dict[str, Any]) -> str:
    """QueryPlan을 간단한 요약 문자열로 변환"""
    parts = []

    entity = query_plan.get("entity", "")
    if entity:
        parts.append(entity)

    # 필터 요약
    filters = query_plan.get("filters", [])
    if filters:
        filter_strs = [f"{f.get('field')} {f.get('operator')} {f.get('value')}" for f in filters]
        parts.append(f"filters:[{', '.join(filter_strs)}]")

    # limit
    if query_plan.get("limit"):
        parts.append(f"limit:{query_plan['limit']}")

    # orderBy
    order_by = query_plan.get("orderBy", [])
    if order_by:
        order_strs = [f"{o.get('field')} {o.get('direction', 'asc')}" for o in order_by]
        parts.append(f"orderBy:[{', '.join(order_strs)}]")

    return ", ".join(parts) if parts else "[쿼리 없음]"


def build_conversation_context(history: List["ChatMessageItem"]) -> str:
    """이전 대화를 프롬프트용 텍스트로 변환 (구조화된 결과 현황 포함)"""
    if not history:
        return ""

    context = "## 이전 대화 컨텍스트\n\n"

    # ============================================
    # 조회 결과 현황 (구조화된 테이블 형식)
    # ============================================
    result_messages = []
    for i, msg in enumerate(history):
        if msg.role == "assistant" and msg.queryResult:
            entity = msg.queryPlan.get("entity", "unknown") if msg.queryPlan else "unknown"
            count = msg.queryResult.get("totalCount", 0)

            # 필터 정보 추출
            filters = msg.queryPlan.get("filters", []) if msg.queryPlan else []
            filter_desc = "-"
            if filters:
                filter_strs = [f"{f.get('field')}={f.get('value')}" for f in filters[:2]]
                filter_desc = ", ".join(filter_strs)

            # 금액 정보 추출
            total_amount = None
            data_obj = msg.queryResult.get("data", {})
            rows = data_obj.get("rows", []) if isinstance(data_obj, dict) else []
            if rows:
                amounts = []
                for row in rows:
                    if isinstance(row, dict):
                        for field in ["amount", "totalAmount", "total_amount", "price"]:
                            if field in row and row[field] is not None:
                                try:
                                    amounts.append(float(row[field]))
                                except (ValueError, TypeError):
                                    pass
                                break
                if amounts:
                    total_amount = sum(amounts)

            # 결과 타입 판단 (테이블 vs 집계)
            result_type = "table"  # 기본값
            if msg.renderSpec and msg.renderSpec.get("type") == "text":
                result_type = "aggregation"

            # 관계 정보 추정
            relation = "최초 조회"
            if len(result_messages) > 0:
                prev = result_messages[-1]
                if prev["entity"] == entity and filters:
                    relation = f"#{prev['index']}에서 필터링"
                elif prev["entity"] != entity:
                    relation = "새로운 엔티티"
                else:
                    relation = "조건 변경"

            result_messages.append({
                "index": i,
                "entity": entity,
                "count": count,
                "filter_desc": filter_desc,
                "total_amount": total_amount,
                "result_type": result_type,
                "relation": relation,
                "is_latest": False
            })

    if result_messages:
        result_messages[-1]["is_latest"] = True  # 마지막이 직전 결과

        # 구조화된 테이블 형식
        context += "### 📊 조회 결과 현황\n"
        context += "| # | 엔티티 | 건수 | 조건 | 금액 | 타입 | 관계 |\n"
        context += "|---|--------|------|------|------|------|------|\n"

        for r in result_messages:
            marker = "👉" if r["is_latest"] else ""
            amount_str = f"${r['total_amount']:,.0f}" if r['total_amount'] else "-"
            context += f"| {marker}{r['index']} | {r['entity']} | {r['count']} | {r['filter_desc']} | {amount_str} | {r['result_type']} | {r['relation']} |\n"

        context += "\n"

        # ============================================
        # 결과 관계 분석 (LLM이 이해하기 쉽게)
        # ============================================
        context += "### 결과 관계 분석\n"
        entities = {}
        for r in result_messages:
            if r["entity"] not in entities:
                entities[r["entity"]] = []
            entities[r["entity"]].append(r)

        for entity, results in entities.items():
            if len(results) > 1:
                context += f"- **{entity}**: {len(results)}개 결과 (조건이 다름)\n"
                for r in results[1:]:
                    context += f"  - 결과 #{r['index']}은 #{result_messages[0]['index']}에서 파생됨\n"
            else:
                context += f"- **{entity}**: 1개 결과\n"

        context += "\n"

        # ============================================
        # 계산에 사용할 데이터 (명시적)
        # ============================================
        latest = result_messages[-1]
        context += "### 📌 현재 작업 대상 (직전 결과)\n"
        context += f"- **엔티티**: {latest['entity']}\n"
        context += f"- **건수**: {latest['count']}건\n"
        context += f"- **타입**: {latest['result_type']} ({'목록 데이터' if latest['result_type'] == 'table' else '집계 결과'})\n"
        if latest['total_amount']:
            context += f"- **금액 합계**: ${latest['total_amount']:,.0f}\n"
        if latest['filter_desc'] != "-":
            context += f"- **적용된 필터**: {latest['filter_desc']}\n"

        context += "\n"

        # 다중 결과 경고
        if len(result_messages) > 1:
            entity_set = set(r["entity"] for r in result_messages)
            if len(entity_set) > 1:
                context += f"⚠️ **주의**: 다른 종류의 결과가 {len(result_messages)}개 있습니다 ({', '.join(entity_set)})\n"
                context += "→ 참조 표현 없으면 어떤 결과를 대상으로 하는지 불명확할 수 있음\n\n"

    # ============================================
    # 대화 히스토리 (최근 5개)
    # ============================================
    context += "### 대화 히스토리\n"
    for msg in history[-5:]:
        if msg.role == 'user':
            context += f"**사용자**: {msg.content}\n"
        else:
            # queryPlan 요약 포함
            if msg.queryPlan:
                plan_summary = summarize_query_plan(msg.queryPlan)
                context += f"**어시스턴트**: [쿼리: {plan_summary}]\n"
            else:
                context += f"**어시스턴트**: [결과 표시됨]\n"

            # 집계 결과값 포함 (중요: 후속 계산용)
            if msg.renderSpec and msg.renderSpec.get("type") == "text":
                text_content = msg.renderSpec.get("text", {}).get("content", "")
                if text_content and ("합계" in text_content or "$" in text_content or "원" in text_content):
                    context += f"  → **집계 결과**: {text_content}\n"

    # ============================================
    # 후속 질문 처리 규칙 (강화)
    # ============================================
    context += "\n### 후속 질문 처리 규칙\n"
    context += "1. **참조 표현 있음** ('이중에', '여기서', '직전', '방금', '아까 그거') → 직전 결과 사용\n"
    context += "2. **참조 표현 없음** + 다중 결과 → 문맥상 명확하지 않으면 clarification 고려\n"
    context += "3. **직전 결과 타입 확인**:\n"
    context += "   - 테이블(목록) + '합산' → aggregate_local\n"
    context += "   - 집계결과 + '수수료 적용' → direct_answer\n"
    context += "   - 집계결과 + '필터링' → query_needed (집계 결과는 필터 불가)\n"
    context += "4. **엔티티 유지**: 후속 질문에서 다른 엔티티로 변경하려면 명시적 표현 필요\n"

    return context


def get_previous_query_plan(history: List["ChatMessageItem"]) -> Optional[Dict[str, Any]]:
    """이전 대화에서 마지막 queryPlan 추출"""
    if not history:
        return None

    # 역순으로 탐색하여 가장 최근 assistant의 queryPlan 찾기
    for msg in reversed(history):
        if msg.role == 'assistant' and msg.queryPlan:
            return msg.queryPlan
    return None


def extract_previous_results(history: List["ChatMessageItem"]) -> List[Dict[str, Any]]:
    """이전 대화에서 조회/집계 결과 요약 추출 (Intent Classification용)

    실제 데이터 값도 추출하여 LLM이 계산할 수 있도록 함
    """
    results = []
    if not history:
        return results

    for i, msg in enumerate(history):
        if msg.role == "assistant":
            result_info = {
                "index": i,
                "entity": None,
                "count": 0,
                "aggregation": None,
                "data_summary": None,  # 실제 데이터 요약
                "total_amount": None   # 금액 합계 (있는 경우)
            }

            # QueryResult가 있으면 조회 결과
            if msg.queryResult:
                logger.info(f"[extract_previous_results] msg #{i} has queryResult with keys: {list(msg.queryResult.keys())}")
                result_info["count"] = msg.queryResult.get("totalCount", 0)
                if msg.queryPlan:
                    result_info["entity"] = msg.queryPlan.get("entity", "unknown")

                # 실제 데이터에서 금액 합계 추출
                data_obj = msg.queryResult.get("data", {})
                # data is an object with 'rows' property according to query-result.schema.json
                rows = data_obj.get("rows", []) if isinstance(data_obj, dict) else []
                logger.info(f"[extract_previous_results] msg #{i} rows length: {len(rows) if rows else 0}")

                if rows:
                    # amount 필드가 있으면 합계 계산
                    amounts = []
                    for row_idx, row in enumerate(rows):
                        if isinstance(row, dict):
                            logger.info(f"[extract_previous_results] msg #{i} row #{row_idx} keys: {list(row.keys())}")
                            # amount, totalAmount, 금액 등 다양한 필드명 체크
                            for field in ["amount", "totalAmount", "total_amount", "price", "금액"]:
                                if field in row and row[field] is not None:
                                    try:
                                        amounts.append(float(row[field]))
                                        logger.info(f"[extract_previous_results] msg #{i} row #{row_idx} found {field}={row[field]}")
                                    except (ValueError, TypeError) as e:
                                        logger.info(f"[extract_previous_results] msg #{i} row #{row_idx} error converting {field}: {e}")
                                        pass
                                    break
                        else:
                            logger.info(f"[extract_previous_results] msg #{i} row #{row_idx} is not a dict: {type(row)}")

                    if amounts:
                        result_info["total_amount"] = sum(amounts)
                        result_info["data_summary"] = f"금액 합계: ${result_info['total_amount']:,.0f} ({len(amounts)}건)"
                        logger.info(f"[extract_previous_results] msg #{i} extracted total_amount: ${result_info['total_amount']:,.0f} from {len(amounts)} amounts")
                    else:
                        logger.info(f"[extract_previous_results] msg #{i} no amounts found in {len(rows)} rows")

            # RenderSpec이 text 타입이면 집계 결과일 수 있음
            if msg.renderSpec and msg.renderSpec.get("type") == "text":
                text_content = msg.renderSpec.get("text", {}).get("content", "")
                if text_content:
                    result_info["aggregation"] = text_content

                    # 텍스트에서 금액 추출 (우선순위: 괄호 안 전체 금액 > 축약 금액)
                    if result_info["total_amount"] is None:
                        # 1순위: 괄호 안의 전체 금액 (예: "$2.88M ($2,878,000)" → 2878000)
                        full_amount_match = re.search(r'\(\$?([\d,]+)\)', text_content)
                        if full_amount_match:
                            try:
                                result_info["total_amount"] = float(full_amount_match.group(1).replace(',', ''))
                                logger.info(f"[extract_previous_results] Extracted full amount from parens: ${result_info['total_amount']:,.0f}")
                            except ValueError:
                                pass

                        # 2순위: M/K 접미사 처리 (예: "$2.88M" → 2880000)
                        if result_info["total_amount"] is None:
                            abbrev_match = re.search(r'\$?([\d,]+(?:\.\d+)?)\s*([MK])', text_content)
                            if abbrev_match:
                                try:
                                    value = float(abbrev_match.group(1).replace(',', ''))
                                    suffix = abbrev_match.group(2)
                                    if suffix == 'M':
                                        value *= 1_000_000
                                    elif suffix == 'K':
                                        value *= 1_000
                                    result_info["total_amount"] = value
                                    logger.info(f"[extract_previous_results] Extracted abbreviated amount: ${result_info['total_amount']:,.0f}")
                                except ValueError:
                                    pass

                        # 3순위: 일반 금액 (예: "$1,234,567")
                        if result_info["total_amount"] is None:
                            simple_match = re.search(r'\$?([\d,]+(?:\.\d+)?)', text_content)
                            if simple_match:
                                try:
                                    result_info["total_amount"] = float(simple_match.group(1).replace(',', ''))
                                    logger.info(f"[extract_previous_results] Extracted simple amount: ${result_info['total_amount']:,.0f}")
                                except ValueError:
                                    pass

            # 조회 결과나 집계 결과가 있으면 추가
            if result_info["count"] > 0 or result_info["aggregation"]:
                results.append(result_info)
                logger.info(f"[extract_previous_results] Added result #{len(results)}: entity={result_info['entity']}, count={result_info['count']}, total_amount={result_info['total_amount']}")

    logger.info(f"[extract_previous_results] Total results extracted: {len(results)}")
    return results


def merge_filters(previous_plan: Dict[str, Any], new_plan: Dict[str, Any]) -> Dict[str, Any]:
    """이전 필터와 새 필터를 병합"""
    if not previous_plan:
        return new_plan

    # clarification 요청이면 병합하지 않음
    if new_plan.get("needs_clarification"):
        return new_plan

    # 이전 필터 가져오기
    prev_filters = previous_plan.get("filters", [])
    new_filters = new_plan.get("filters", [])

    # 새 필터의 필드명 목록
    new_filter_fields = {f.get("field") for f in new_filters}

    # 이전 필터 중 새 필터에 없는 것만 병합 (중복 필드 방지)
    merged_filters = list(new_filters)  # 새 필터 우선
    for prev_filter in prev_filters:
        if prev_filter.get("field") not in new_filter_fields:
            merged_filters.append(prev_filter)

    # 병합된 결과
    merged_plan = dict(new_plan)
    if merged_filters:
        merged_plan["filters"] = merged_filters

    # 이전 entity 유지 (새 plan에 entity가 없으면)
    if not merged_plan.get("entity") and previous_plan.get("entity"):
        merged_plan["entity"] = previous_plan["entity"]

    # 이전 limit 유지 (새 plan이 기본값 10이면)
    if merged_plan.get("limit") == 10 and previous_plan.get("limit"):
        merged_plan["limit"] = previous_plan["limit"]

    return merged_plan

from app.services.query_planner import get_query_planner, IntentType
from app.services.render_composer import get_render_composer
from app.services.rag_service import get_rag_service

# Text-to-SQL 모드용 import (조건부)
if ENABLE_TEXT_TO_SQL:
    from app.services.text_to_sql import get_text_to_sql_service, extract_where_conditions

logger = logging.getLogger(__name__)
logger.info(f"Text-to-SQL mode: {'ENABLED' if ENABLE_TEXT_TO_SQL else 'DISABLED'}")

router = APIRouter()

# Configuration
CORE_API_URL = os.getenv("CORE_API_URL", "http://localhost:8080")
ENABLE_QUERY_PLAN_VALIDATION = os.getenv("ENABLE_QUERY_PLAN_VALIDATION", "true").lower() == "true"


class ChatMessageItem(BaseModel):
    """대화 메시지 아이템"""
    id: str
    role: str  # 'user' | 'assistant'
    content: str
    timestamp: str
    status: Optional[str] = None
    renderSpec: Optional[Dict[str, Any]] = None
    queryResult: Optional[Dict[str, Any]] = None
    queryPlan: Optional[Dict[str, Any]] = None  # 이전 쿼리 조건 저장용


class ChatRequest(BaseModel):
    """채팅 요청"""
    message: str
    conversation_id: Optional[str] = None
    session_id: Optional[str] = Field(default=None, alias="sessionId")
    conversation_history: Optional[List[ChatMessageItem]] = Field(default=None, alias="conversationHistory")

    class Config:
        populate_by_name = True


class ChatResponse(BaseModel):
    """채팅 응답 - UI 타입과 일치"""
    request_id: str = Field(alias="requestId")
    render_spec: Dict[str, Any] = Field(alias="renderSpec")
    query_result: Optional[Dict[str, Any]] = Field(default=None, alias="queryResult")  # filter_local 시 None 가능
    query_plan: Dict[str, Any] = Field(alias="queryPlan")  # 이번 쿼리 조건 (후속 질문용)
    ai_message: Optional[str] = Field(default=None, alias="aiMessage")
    timestamp: str

    class Config:
        populate_by_name = True


@router.post("/chat", response_model=ChatResponse, response_model_by_alias=True)
async def chat(request: ChatRequest):
    """
    Step 6: LangChain 기반 자연어 처리

    Flow (기존 QueryPlan 모드):
    1. 사용자 메시지 수신
    2. QueryPlannerService로 자연어 → QueryPlan 변환
    3. Core API 호출
    4. RenderComposerService로 QueryResult → RenderSpec 변환
    5. RenderSpec 반환

    Flow (Text-to-SQL 모드, SQL_ENABLE_TEXT_TO_SQL=true):
    1. 사용자 메시지 수신
    2. TextToSqlService로 자연어 → SQL 변환
    3. SQL 검증 (SqlValidator)
    4. 읽기 전용 DB에서 직접 실행
    5. 결과를 RenderSpec으로 변환
    """
    start_time = datetime.utcnow()
    conversation_id = request.conversation_id or str(uuid.uuid4())
    request_id = f"req-{uuid.uuid4().hex[:8]}"

    logger.info(f"[{request_id}] Received message: {request.message}")

    processing_info = {
        "requestId": request_id,
        "stages": []
    }

    # ========================================
    # Text-to-SQL 모드 분기
    # ========================================
    if ENABLE_TEXT_TO_SQL:
        return await handle_text_to_sql(request, request_id, start_time)

    try:
        # Stage 0: Intent Classification (2단계 분류)
        stage_start = datetime.utcnow()
        query_planner = get_query_planner()

        # 대화 컨텍스트 빌드
        conversation_context = None
        previous_results = []
        if request.conversation_history:
            conversation_context = build_conversation_context(request.conversation_history)
            previous_results = extract_previous_results(request.conversation_history)
            logger.info(f"[{request_id}] Using conversation context with {len(request.conversation_history)} messages")
            logger.info(f"[{request_id}] Found {len(previous_results)} previous results for intent classification")

        # 1단계: Intent 분류 (가벼운 모델로 빠르게)
        intent_result = await query_planner.classify_intent(
            request.message,
            conversation_context or "",
            previous_results
        )
        logger.info(f"[{request_id}] Intent classification: {intent_result.intent.value}, confidence={intent_result.confidence:.2f}")

        # direct_answer면 바로 응답 반환 (QueryPlan 생성 스킵)
        if intent_result.intent == IntentType.DIRECT_ANSWER and intent_result.direct_answer_text:
            logger.info(f"[{request_id}] Direct answer detected, skipping QueryPlan generation")
            logger.info(f"[{request_id}] Direct answer text: {intent_result.direct_answer_text}")

            return ChatResponse(
                request_id=request_id,
                query_plan={
                    "query_intent": "direct_answer",
                    "requestId": request_id
                },
                query_result=None,
                render_spec={
                    "type": "text",
                    "text": {
                        "content": intent_result.direct_answer_text,
                        "format": "markdown"
                    },
                    "metadata": {
                        "intent": "direct_answer",
                        "confidence": intent_result.confidence,
                        "reasoning": intent_result.reasoning
                    }
                },
                timestamp=datetime.utcnow().isoformat() + "Z"
            )

        processing_info["stages"].append({
            "name": "Intent Classification",
            "duration": (datetime.utcnow() - stage_start).total_seconds() * 1000,
            "result": intent_result.intent.value
        })

        # Stage 1: Natural Language → QueryPlan
        stage_start = datetime.utcnow()

        query_plan = await query_planner.generate_query_plan(
            request.message,
            conversation_context=conversation_context,
            enable_validation=ENABLE_QUERY_PLAN_VALIDATION
        )

        # LLM이 판단한 의도에 따라 필터 병합 결정
        query_intent = query_plan.get("query_intent", "new_query")
        logger.info(f"[{request_id}] Query intent: {query_intent}")

        if query_intent == "refine_previous":
            if request.conversation_history:
                previous_plan = get_previous_query_plan(request.conversation_history)
                if previous_plan:
                    logger.info(f"[{request_id}] Intent: refine_previous, merging with previous filters")
                    logger.info(f"[{request_id}] Previous plan: {previous_plan}")
                    query_plan = merge_filters(previous_plan, query_plan)
                    logger.info(f"[{request_id}] Merged plan: {query_plan}")
        elif query_intent == "filter_local":
            # 클라이언트 사이드 필터링: Core API 호출 없이 필터 조건만 반환
            logger.info(f"[{request_id}] Intent: filter_local, client-side filtering")

            # entity가 없으면 이전 queryPlan에서 상속
            if not query_plan.get("entity") and request.conversation_history:
                previous_plan = get_previous_query_plan(request.conversation_history)
                if previous_plan and previous_plan.get("entity"):
                    query_plan["entity"] = previous_plan["entity"]
                    logger.info(f"[{request_id}] Inherited entity from previous plan: {query_plan['entity']}")

            # 이전 결과가 있는 메시지들 찾기
            result_messages = []
            if request.conversation_history:
                logger.info(f"[{request_id}] Checking {len(request.conversation_history)} messages in history")
                for i, msg in enumerate(request.conversation_history):
                    has_query_result = msg.queryResult is not None
                    logger.info(f"[{request_id}] Message {i}: role={msg.role}, hasQueryResult={has_query_result}")
                    if msg.role == "assistant" and msg.queryResult:
                        result_messages.append((i, msg))

            logger.info(f"[{request_id}] Found {len(result_messages)} result messages")

            # 1단계: LLM이 모호하다고 판단했는지 확인
            needs_result_clarification = query_plan.get("needs_result_clarification", False)
            logger.info(f"[{request_id}] 1st stage LLM decision: needs_result_clarification={needs_result_clarification}")

            # 2단계: 다중 결과 + 1단계가 False면 상위 모델로 재판단
            if len(result_messages) > 1 and not needs_result_clarification:
                logger.info(f"[{request_id}] Multiple results but 1st stage said no clarification, invoking 2nd stage check...")

                # 결과 요약 생성
                result_summaries = []
                for msg_idx, msg in result_messages:
                    entity = msg.queryPlan.get("entity", "unknown") if msg.queryPlan else "unknown"
                    count = 0
                    if msg.queryResult:
                        if isinstance(msg.queryResult, dict):
                            count = msg.queryResult.get("totalCount", msg.queryResult.get("metadata", {}).get("rowsReturned", 0))
                    filters_str = ""
                    if msg.queryPlan and msg.queryPlan.get("filters"):
                        filters_str = ", ".join([f"{f.get('field')}={f.get('value')}" for f in msg.queryPlan.get("filters", [])[:2]])
                    result_summaries.append({"entity": entity, "count": count, "filters": filters_str})

                # 2단계 LLM 판단 호출
                needs_result_clarification = await query_planner.check_clarification_needed(
                    user_message=request.message,
                    result_summaries=result_summaries,
                    query_intent=query_intent
                )
                logger.info(f"[{request_id}] 2nd stage LLM decision: needs_result_clarification={needs_result_clarification}")

            if len(result_messages) > 1 and needs_result_clarification:
                # 다중 결과 + LLM이 모호하다고 판단: clarification 요청
                recent_results = result_messages[-5:]  # 최근 5개만
                options = []
                indices = []

                for idx, (msg_idx, msg) in enumerate(reversed(recent_results)):
                    # 결과 요약 라벨 생성
                    entity = msg.queryPlan.get("entity", "데이터") if msg.queryPlan else "데이터"
                    count = "?"
                    if msg.queryResult:
                        if isinstance(msg.queryResult, dict):
                            count = msg.queryResult.get("totalCount", msg.queryResult.get("metadata", {}).get("rowsReturned", "?"))
                        elif hasattr(msg.queryResult, "totalCount"):
                            count = msg.queryResult.totalCount
                    time_str = msg.timestamp[-8:-3] if msg.timestamp and len(msg.timestamp) >= 8 else ""

                    label = f"직전: {entity} {count}건 ({time_str})" if idx == 0 else f"{entity} {count}건 ({time_str})"
                    options.append(label)
                    indices.append(msg_idx)

                logger.info(f"[{request_id}] Multiple results found, requesting clarification")

                clarification_render_spec = {
                    "type": "clarification",
                    "clarification": {
                        "question": "어떤 조회 결과를 필터링할까요?",
                        "options": options
                    },
                    "metadata": {
                        "requestId": request_id,
                        "targetResultIndices": indices,
                        "pendingFilters": query_plan.get("filters", []),
                        "generatedAt": datetime.utcnow().isoformat() + "Z"
                    }
                }

                return ChatResponse(
                    request_id=request_id,
                    render_spec=clarification_render_spec,
                    query_result=None,
                    query_plan={**query_plan, "needs_clarification": True, "requestId": request_id},
                    ai_message="어떤 조회 결과를 필터링할까요?",
                    timestamp=datetime.utcnow().isoformat() + "Z"
                )

            # 결과가 1개 이하: 클라이언트에서 필터링하도록 응답
            target_index = result_messages[-1][0] if result_messages else -1
            logger.info(f"[{request_id}] Single result, returning filter_local response (target: {target_index})")

            filter_local_render_spec = {
                "type": "filter_local",
                "filter": query_plan.get("filters", []),
                "targetResultIndex": target_index,
                "metadata": {
                    "requestId": request_id,
                    "generatedAt": datetime.utcnow().isoformat() + "Z"
                }
            }

            return ChatResponse(
                request_id=request_id,
                render_spec=filter_local_render_spec,
                query_result=None,
                query_plan={**query_plan, "requestId": request_id},
                ai_message="이전 결과에서 필터링합니다.",
                timestamp=datetime.utcnow().isoformat() + "Z"
            )
        elif query_intent == "aggregate_local":
            # 클라이언트 사이드 집계: 이전 결과에서 집계
            logger.info(f"[{request_id}] Intent: aggregate_local, client-side aggregation")

            # entity가 없으면 이전 queryPlan에서 상속
            if not query_plan.get("entity") and request.conversation_history:
                previous_plan = get_previous_query_plan(request.conversation_history)
                if previous_plan and previous_plan.get("entity"):
                    query_plan["entity"] = previous_plan["entity"]
                    logger.info(f"[{request_id}] Inherited entity from previous plan: {query_plan['entity']}")

            # 이전 결과가 있는 메시지들 찾기
            result_messages = []
            if request.conversation_history:
                logger.info(f"[{request_id}] Checking {len(request.conversation_history)} messages in history")
                for i, msg in enumerate(request.conversation_history):
                    has_query_result = msg.queryResult is not None
                    if msg.role == "assistant" and msg.queryResult:
                        result_messages.append((i, msg))

            logger.info(f"[{request_id}] Found {len(result_messages)} result messages for aggregation")

            # 집계 정보 추출
            aggregations = query_plan.get("aggregations", [])
            if not aggregations:
                # 기본 집계: sum(amount)
                aggregations = [{"function": "sum", "field": "amount", "alias": "totalAmount", "displayLabel": "결제 금액 합계", "currency": "USD"}]

            # 1단계: LLM이 모호하다고 판단했는지 확인
            needs_result_clarification = query_plan.get("needs_result_clarification", False)
            logger.info(f"[{request_id}] 1st stage LLM decision (aggregate): needs_result_clarification={needs_result_clarification}")

            # 2단계: 다중 결과 + 1단계가 False면 상위 모델로 재판단
            if len(result_messages) > 1 and not needs_result_clarification:
                logger.info(f"[{request_id}] Multiple results but 1st stage said no clarification, invoking 2nd stage check...")

                # 결과 요약 생성
                result_summaries = []
                for msg_idx, msg in result_messages:
                    entity = msg.queryPlan.get("entity", "unknown") if msg.queryPlan else "unknown"
                    count = 0
                    if msg.queryResult:
                        if isinstance(msg.queryResult, dict):
                            count = msg.queryResult.get("totalCount", msg.queryResult.get("metadata", {}).get("rowsReturned", 0))
                    filters_str = ""
                    if msg.queryPlan and msg.queryPlan.get("filters"):
                        filters_str = ", ".join([f"{f.get('field')}={f.get('value')}" for f in msg.queryPlan.get("filters", [])[:2]])
                    result_summaries.append({"entity": entity, "count": count, "filters": filters_str})

                # 2단계 LLM 판단 호출
                needs_result_clarification = await query_planner.check_clarification_needed(
                    user_message=request.message,
                    result_summaries=result_summaries,
                    query_intent=query_intent
                )
                logger.info(f"[{request_id}] 2nd stage LLM decision (aggregate): needs_result_clarification={needs_result_clarification}")

            if len(result_messages) > 1 and needs_result_clarification:
                # 다중 결과 + LLM이 모호하다고 판단: clarification 요청
                recent_results = result_messages[-5:]  # 최근 5개만
                options = []
                indices = []

                for idx, (msg_idx, msg) in enumerate(reversed(recent_results)):
                    entity = msg.queryPlan.get("entity", "데이터") if msg.queryPlan else "데이터"
                    count = "?"
                    if msg.queryResult:
                        if isinstance(msg.queryResult, dict):
                            count = msg.queryResult.get("totalCount", msg.queryResult.get("metadata", {}).get("rowsReturned", "?"))
                    time_str = msg.timestamp[-8:-3] if msg.timestamp and len(msg.timestamp) >= 8 else ""

                    label = f"직전: {entity} {count}건 ({time_str})" if idx == 0 else f"{entity} {count}건 ({time_str})"
                    options.append(label)
                    indices.append(msg_idx)

                logger.info(f"[{request_id}] Multiple results found, requesting clarification for aggregation")

                clarification_render_spec = {
                    "type": "clarification",
                    "clarification": {
                        "question": "어떤 데이터를 기준으로 집계할까요?",
                        "options": options
                    },
                    "metadata": {
                        "requestId": request_id,
                        "targetResultIndices": indices,
                        "pendingAggregations": aggregations,
                        "aggregationType": "aggregate_local",
                        "generatedAt": datetime.utcnow().isoformat() + "Z"
                    }
                }

                return ChatResponse(
                    request_id=request_id,
                    render_spec=clarification_render_spec,
                    query_result=None,
                    query_plan={**query_plan, "needs_clarification": True, "requestId": request_id},
                    ai_message="어떤 데이터를 기준으로 집계할까요?",
                    timestamp=datetime.utcnow().isoformat() + "Z"
                )

            # 결과가 1개 이하: 클라이언트에서 집계하도록 응답
            target_index = result_messages[-1][0] if result_messages else -1
            logger.info(f"[{request_id}] Single result, returning aggregate_local response (target: {target_index})")

            aggregate_local_render_spec = {
                "type": "aggregate_local",
                "aggregations": aggregations,
                "targetResultIndex": target_index,
                "metadata": {
                    "requestId": request_id,
                    "generatedAt": datetime.utcnow().isoformat() + "Z"
                }
            }

            return ChatResponse(
                request_id=request_id,
                render_spec=aggregate_local_render_spec,
                query_result=None,
                query_plan={**query_plan, "requestId": request_id},
                ai_message="이전 결과에서 집계합니다.",
                timestamp=datetime.utcnow().isoformat() + "Z"
            )
        elif query_intent == "direct_answer":
            # LLM이 직접 답변: DB 조회 없이 텍스트 응답
            direct_answer = query_plan.get("direct_answer", "")
            logger.info(f"[{request_id}] Intent: direct_answer, returning LLM response")

            if not direct_answer:
                direct_answer = "죄송합니다. 답변을 생성하지 못했습니다."

            direct_answer_render_spec = {
                "type": "text",
                "title": "분석 결과",
                "text": {
                    "content": direct_answer,
                    "format": "markdown"
                },
                "metadata": {
                    "requestId": request_id,
                    "generatedAt": datetime.utcnow().isoformat() + "Z"
                }
            }

            return ChatResponse(
                request_id=request_id,
                render_spec=direct_answer_render_spec,
                query_result=None,
                query_plan={**query_plan, "requestId": request_id},
                ai_message=direct_answer,
                timestamp=datetime.utcnow().isoformat() + "Z"
            )
        else:
            logger.info(f"[{request_id}] Intent: new_query, no filter merge")

        query_plan["requestId"] = request_id

        processing_info["stages"].append({
            "name": "query_plan_generation",
            "durationMs": int((datetime.utcnow() - stage_start).total_seconds() * 1000),
            "status": "success"
        })

        logger.info(f"[{request_id}] Final QueryPlan: {query_plan}")

        # Clarification 필요 시 쿼리 실행 없이 대화형 질문 반환
        if query_plan.get("needs_clarification"):
            question = query_plan.get("clarification_question", "어떤 데이터를 조회하시겠습니까?")
            logger.info(f"[{request_id}] Clarification needed: {question}")

            # 대화형 텍스트로 응답 (버튼 없이)
            clarification_render_spec = {
                "type": "text",
                "title": "추가 정보 필요",
                "text": {
                    "content": question,
                    "format": "plain"
                },
                "metadata": {
                    "requestId": request_id,
                    "generatedAt": datetime.utcnow().isoformat() + "Z"
                }
            }

            clarification_query_result = {
                "requestId": request_id,
                "status": "pending",
                "data": {"rows": [], "aggregations": {}},
                "metadata": {
                    "executionTimeMs": int((datetime.utcnow() - start_time).total_seconds() * 1000),
                    "rowsReturned": 0,
                    "dataSource": "clarification_required"
                }
            }

            return ChatResponse(
                request_id=request_id,
                render_spec=clarification_render_spec,
                query_result=clarification_query_result,
                query_plan=query_plan,
                ai_message=question,
                timestamp=datetime.utcnow().isoformat() + "Z"
            )

        # Stage 2: Call Core API
        stage_start = datetime.utcnow()
        query_result = await call_core_api(query_plan)

        processing_info["stages"].append({
            "name": "core_api_call",
            "durationMs": int((datetime.utcnow() - stage_start).total_seconds() * 1000),
            "status": "success" if query_result.get("status") == "success" else "error"
        })

        logger.info(f"[{request_id}] Core API response status: {query_result.get('status')}")

        # Stage 3: QueryResult → RenderSpec
        stage_start = datetime.utcnow()
        render_composer = get_render_composer()
        render_spec = render_composer.compose(query_result, query_plan, request.message)

        processing_info["stages"].append({
            "name": "render_spec_composition",
            "durationMs": int((datetime.utcnow() - stage_start).total_seconds() * 1000),
            "status": "success"
        })

        # Calculate total processing time
        total_time = int((datetime.utcnow() - start_time).total_seconds() * 1000)
        processing_info["totalDurationMs"] = total_time

        logger.info(f"[{request_id}] Completed in {total_time}ms")

        return ChatResponse(
            request_id=request_id,
            render_spec=render_spec,
            query_result=query_result,
            query_plan=query_plan,  # 후속 질문에서 이전 쿼리 조건 참조용
            ai_message=f"'{request.message}'에 대한 결과입니다.",
            timestamp=datetime.utcnow().isoformat() + "Z"
        )

    except Exception as e:
        logger.error(f"[{request_id}] Error: {e}", exc_info=True)

        total_time = int((datetime.utcnow() - start_time).total_seconds() * 1000)
        processing_info["totalDurationMs"] = total_time
        processing_info["error"] = str(e)

        # 에러 발생 시에도 RenderSpec 반환 (에러 메시지 표시용)
        error_render_spec = {
            "type": "text",
            "title": "처리 중 오류 발생",
            "text": {
                "content": f"## 요청 처리 중 오류가 발생했습니다\n\n"
                          f"**요청**: {request.message}\n\n"
                          f"**오류**: {str(e)}\n\n"
                          f"잠시 후 다시 시도해주세요.",
                "format": "markdown",
                "sections": [
                    {
                        "type": "error",
                        "title": "오류 정보",
                        "content": str(e)
                    }
                ]
            },
            "metadata": {
                "requestId": request_id,
                "generatedAt": datetime.utcnow().isoformat() + "Z"
            }
        }

        # 에러 시 빈 QueryResult 반환
        error_query_result = {
            "requestId": request_id,
            "status": "error",
            "data": {"rows": [], "aggregations": {}},
            "metadata": {
                "executionTimeMs": total_time,
                "rowsReturned": 0,
                "totalRows": 0,
                "dataSource": "error"
            },
            "error": {
                "code": "PROCESSING_ERROR",
                "message": str(e)
            }
        }

        # 에러 시 빈 QueryPlan 반환
        error_query_plan = {
            "entity": "",
            "operation": "list"
        }

        return ChatResponse(
            request_id=request_id,
            render_spec=error_render_spec,
            query_result=error_query_result,
            query_plan=error_query_plan,
            ai_message=f"요청 처리 중 오류가 발생했습니다: {str(e)}",
            timestamp=datetime.utcnow().isoformat() + "Z"
        )


async def call_core_api(query_plan: Dict[str, Any]) -> Dict[str, Any]:
    """Core API 호출"""
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(
                f"{CORE_API_URL}/api/v1/query/start",
                json=query_plan
            )

            # HTTP 에러가 아닌 비즈니스 에러도 처리
            if response.status_code >= 400:
                error_body = response.json() if response.headers.get("content-type", "").startswith("application/json") else {}
                logger.warning(f"Core API returned {response.status_code}: {error_body}")
                return error_body if error_body else {
                    "status": "error",
                    "error": {
                        "code": f"HTTP_{response.status_code}",
                        "message": response.text
                    }
                }

            return response.json()

    except httpx.TimeoutException:
        logger.error("Core API timeout")
        return {
            "status": "error",
            "error": {
                "code": "TIMEOUT",
                "message": "Core API 요청 시간이 초과되었습니다."
            }
        }
    except httpx.HTTPError as e:
        logger.error(f"Core API HTTP error: {e}")
        return {
            "status": "error",
            "error": {
                "code": "CONNECTION_ERROR",
                "message": f"Core API 연결 오류: {str(e)}"
            }
        }


@router.get("/chat/test")
async def test_core_api():
    """Core API 연결 테스트"""
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            response = await client.get(f"{CORE_API_URL}/api/v1/query/health")
            response.raise_for_status()
            return {
                "core_api_status": "reachable",
                "core_api_response": response.json()
            }
    except httpx.HTTPError as e:
        return {
            "core_api_status": "unreachable",
            "error": str(e)
        }


@router.get("/chat/config")
async def get_config():
    """현재 설정 확인 (디버깅용)"""
    return {
        "core_api_url": CORE_API_URL,
        "openai_api_key_set": bool(os.getenv("OPENAI_API_KEY")),
        "anthropic_api_key_set": bool(os.getenv("ANTHROPIC_API_KEY")),
        "rag_enabled": os.getenv("RAG_ENABLED", "true").lower() == "true",
        "database_url_set": bool(os.getenv("DATABASE_URL")),
        "query_plan_validation": {
            "enabled": ENABLE_QUERY_PLAN_VALIDATION,
            "validator_provider": os.getenv("VALIDATOR_LLM_PROVIDER", "openai"),
            "validator_model": os.getenv("VALIDATOR_LLM_MODEL", "gpt-4o-mini"),
            "quality_threshold": float(os.getenv("VALIDATOR_QUALITY_THRESHOLD", "0.8")),
            "use_llm_validation": os.getenv("VALIDATOR_USE_LLM", "true").lower() == "true"
        },
        "step": "8-query-plan-validation"
    }


@router.get("/chat/rag/status")
async def get_rag_status():
    """RAG 서비스 상태 확인"""
    try:
        rag_service = get_rag_service()
        doc_counts = await rag_service.get_document_count()

        return {
            "status": "available",
            "rag_enabled": os.getenv("RAG_ENABLED", "true").lower() == "true",
            "openai_configured": bool(os.getenv("OPENAI_API_KEY")),
            "document_counts": doc_counts,
            "total_documents": sum(doc_counts.values()) if doc_counts else 0
        }
    except Exception as e:
        return {
            "status": "unavailable",
            "error": str(e)
        }


@router.post("/chat/rag/search")
async def search_documents(query: str, k: int = 3):
    """RAG 문서 검색 테스트"""
    try:
        rag_service = get_rag_service()
        documents = await rag_service.search_docs(query=query, k=k)

        return {
            "query": query,
            "count": len(documents),
            "documents": [
                {
                    "id": doc.id,
                    "doc_type": doc.doc_type,
                    "title": doc.title,
                    "content": doc.content[:200] + "..." if len(doc.content) > 200 else doc.content,
                    "similarity": doc.similarity
                }
                for doc in documents
            ]
        }
    except Exception as e:
        return {
            "query": query,
            "error": str(e)
        }


# ============================================
# 대용량 데이터 다운로드 엔드포인트
# ============================================

class DownloadRequest(BaseModel):
    """다운로드 요청"""
    sql: str
    format: str = "csv"  # csv 또는 excel


@router.post("/chat/download")
async def download_query_result(request: DownloadRequest):
    """
    대용량 쿼리 결과 다운로드

    - SQL 재검증 후 실행 (LIMIT 없이)
    - Streaming 응답으로 메모리 효율화
    """
    if not ENABLE_TEXT_TO_SQL:
        raise HTTPException(400, "Text-to-SQL mode is not enabled")

    from app.services.sql_validator import get_sql_validator
    from app.services.text_to_sql import get_text_to_sql_service

    # SQL 검증 (보안)
    validator = get_sql_validator()
    validation = validator.validate(request.sql)

    if not validation.is_valid:
        raise HTTPException(400, f"Invalid SQL: {', '.join(validation.issues)}")

    # LIMIT 제거 (전체 데이터 다운로드)
    unlimited_sql = re.sub(r'\bLIMIT\s+\d+', '', validation.sanitized_sql, flags=re.IGNORECASE)
    unlimited_sql = re.sub(r'\bOFFSET\s+\d+', '', unlimited_sql, flags=re.IGNORECASE)

    logger.info(f"Download request - Original SQL: {request.sql[:100]}...")
    logger.info(f"Download request - Unlimited SQL: {unlimited_sql[:100]}...")

    text_to_sql = get_text_to_sql_service()

    def generate_csv() -> Generator[str, None, None]:
        """CSV 스트리밍 생성기"""
        import psycopg
        from psycopg.rows import dict_row

        try:
            with text_to_sql._get_readonly_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(unlimited_sql)

                    # 헤더 출력
                    columns = [desc[0] for desc in cur.description]
                    output = io.StringIO()
                    writer = csv.writer(output)
                    writer.writerow(columns)
                    yield output.getvalue()

                    # 데이터 배치 처리 (1000건씩)
                    batch_size = 1000
                    row_count = 0
                    while True:
                        rows = cur.fetchmany(batch_size)
                        if not rows:
                            break

                        output = io.StringIO()
                        writer = csv.writer(output)
                        for row in rows:
                            # datetime 변환
                            processed_row = []
                            for value in row.values() if hasattr(row, 'values') else row:
                                if hasattr(value, 'isoformat'):
                                    processed_row.append(value.isoformat())
                                else:
                                    processed_row.append(value)
                            writer.writerow(processed_row)
                            row_count += 1

                        yield output.getvalue()

                    logger.info(f"Download completed: {row_count} rows")

        except psycopg.Error as e:
            logger.error(f"Download SQL execution failed: {e}")
            yield f"Error: {str(e)}"

    def generate_excel() -> bytes:
        """Excel 파일 생성 (메모리 내)"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import psycopg

        wb = Workbook()
        ws = wb.active
        ws.title = "Query Result"

        try:
            with text_to_sql._get_readonly_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(unlimited_sql)

                    # 헤더 행 스타일
                    columns = [desc[0] for desc in cur.description]
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

                    for col_idx, col_name in enumerate(columns, 1):
                        cell = ws.cell(row=1, column=col_idx, value=col_name)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")

                    # 데이터 배치 처리 (1000건씩)
                    batch_size = 1000
                    row_num = 2
                    while True:
                        rows = cur.fetchmany(batch_size)
                        if not rows:
                            break

                        for row in rows:
                            values = row.values() if hasattr(row, 'values') else row
                            for col_idx, value in enumerate(values, 1):
                                # datetime을 ISO format 문자열로 변환
                                if hasattr(value, 'isoformat'):
                                    value = value.isoformat()
                                # dict/list (JSONB)를 JSON 문자열로 변환
                                elif isinstance(value, (dict, list)):
                                    import json
                                    value = json.dumps(value, ensure_ascii=False, default=str)
                                ws.cell(row=row_num, column=col_idx, value=value)
                            row_num += 1

                    logger.info(f"Excel generation completed: {row_num - 2} rows")

        except psycopg.Error as e:
            logger.error(f"Excel SQL execution failed: {e}")
            raise HTTPException(500, f"Excel generation failed: {str(e)}")

        # BytesIO로 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # 파일명 생성
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    # format에 따라 분기 처리
    if request.format == "excel":
        try:
            excel_data = generate_excel()
            filename = f"query_result_{timestamp}.xlsx"

            from fastapi.responses import Response
            return Response(
                content=excel_data,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}",
                    "X-Content-Type-Options": "nosniff"
                }
            )
        except Exception as e:
            logger.error(f"Excel download failed: {e}")
            raise HTTPException(500, f"Excel generation failed: {str(e)}")

    # CSV 응답 (기본)
    filename = f"query_result_{timestamp}.csv"
    return StreamingResponse(
        generate_csv(),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "X-Content-Type-Options": "nosniff"
        }
    )


# ============================================
# Text-to-SQL 모드 핸들러
# ============================================

async def handle_text_to_sql(
    request: ChatRequest,
    request_id: str,
    start_time: datetime
) -> ChatResponse:
    """
    Text-to-SQL 모드 처리

    AI가 직접 SQL을 생성하고 읽기 전용 DB에서 실행합니다.
    """
    logger.info(f"[{request_id}] Text-to-SQL mode: processing")

    try:
        text_to_sql = get_text_to_sql_service()

        # 참조 표현 감지 (연속 대화 WHERE 조건 병합용)
        is_refinement, ref_type = detect_reference_expression(request.message)
        if is_refinement:
            logger.info(f"[{request_id}] Reference expression detected (type: {ref_type}), will preserve previous WHERE conditions")

        # 대화 이력 변환 (Text-to-SQL 형식)
        sql_history = build_sql_history(request.conversation_history)

        # SQL 생성 및 실행 (is_refinement 전달)
        result = await text_to_sql.query(
            question=request.message,
            conversation_history=sql_history,
            retry_on_error=True,
            is_refinement=is_refinement
        )

        # 실행 시간 계산
        total_time_ms = int((datetime.utcnow() - start_time).total_seconds() * 1000)

        if result["success"]:
            # 성공: 데이터를 RenderSpec으로 변환
            # LLM 추천 차트 타입 및 인사이트 템플릿 추출
            llm_chart_type = result.get("llmChartType")
            insight_template = result.get("insightTemplate")
            if llm_chart_type:
                logger.info(f"[{request_id}] LLM chart type: {llm_chart_type}")
            if insight_template:
                logger.info(f"[{request_id}] LLM insight template: {insight_template[:50]}...")

            render_spec = compose_sql_render_spec(result, request.message, llm_chart_type, insight_template)

            # 집계 쿼리 메타데이터 추가
            is_aggregation = result.get("isAggregation", False)
            aggregation_context = result.get("aggregationContext")

            query_result = {
                "requestId": request_id,
                "status": "success",
                "data": {
                    "rows": result["data"],
                    "aggregations": {}
                },
                "metadata": {
                    "executionTimeMs": int(result.get("executionTimeMs") or 0),
                    "rowsReturned": result["rowCount"],
                    "totalRows": result["rowCount"],
                    "dataSource": "text_to_sql"
                },
                # 집계 쿼리 정보 추가
                "isAggregation": is_aggregation,
                "aggregationContext": aggregation_context
            }
        else:
            # 실패: 에러 RenderSpec
            render_spec = {
                "type": "text",
                "title": "쿼리 실행 오류",
                "text": {
                    "content": f"## 쿼리 실행 중 오류가 발생했습니다\n\n"
                              f"**질문**: {request.message}\n\n"
                              f"**오류**: {result.get('error', '알 수 없는 오류')}\n\n"
                              f"**생성된 SQL**:\n```sql\n{result.get('sql', 'N/A')}\n```",
                    "format": "markdown"
                },
                "metadata": {
                    "requestId": request_id,
                    "generatedAt": datetime.utcnow().isoformat() + "Z",
                    "mode": "text_to_sql"
                }
            }
            query_result = {
                "requestId": request_id,
                "status": "error",
                "data": {"rows": [], "aggregations": {}},
                "metadata": {
                    "executionTimeMs": total_time_ms,
                    "rowsReturned": 0,
                    "dataSource": "text_to_sql"
                },
                "error": {
                    "code": "SQL_EXECUTION_ERROR",
                    "message": result.get("error", "Unknown error")
                }
            }

        logger.info(f"[{request_id}] Text-to-SQL completed: success={result['success']}, rows={result['rowCount']}")

        return ChatResponse(
            request_id=request_id,
            render_spec=render_spec,
            query_result=query_result,
            query_plan={
                "mode": "text_to_sql",
                "sql": result.get("sql"),
                "requestId": request_id
            },
            ai_message=f"'{request.message}'에 대한 결과입니다." if result["success"] else "쿼리 실행 중 오류가 발생했습니다.",
            timestamp=datetime.utcnow().isoformat() + "Z"
        )

    except Exception as e:
        logger.error(f"[{request_id}] Text-to-SQL error: {e}", exc_info=True)
        total_time_ms = int((datetime.utcnow() - start_time).total_seconds() * 1000)

        error_render_spec = {
            "type": "text",
            "title": "처리 중 오류 발생",
            "text": {
                "content": f"## Text-to-SQL 처리 중 오류가 발생했습니다\n\n"
                          f"**요청**: {request.message}\n\n"
                          f"**오류**: {str(e)}\n\n"
                          f"잠시 후 다시 시도해주세요.",
                "format": "markdown"
            },
            "metadata": {
                "requestId": request_id,
                "generatedAt": datetime.utcnow().isoformat() + "Z",
                "mode": "text_to_sql"
            }
        }

        return ChatResponse(
            request_id=request_id,
            render_spec=error_render_spec,
            query_result={
                "requestId": request_id,
                "status": "error",
                "data": {"rows": [], "aggregations": {}},
                "metadata": {"executionTimeMs": total_time_ms, "rowsReturned": 0}
            },
            query_plan={"mode": "text_to_sql", "error": str(e)},
            ai_message=f"처리 중 오류가 발생했습니다: {str(e)}",
            timestamp=datetime.utcnow().isoformat() + "Z"
        )


def build_sql_history(conversation_history: Optional[List[ChatMessageItem]]) -> List[Dict[str, Any]]:
    """
    대화 이력을 Text-to-SQL 형식으로 변환

    대화 기반 맥락 처리를 위해 다음 정보를 포함:
    - role: 메시지 역할 (user/assistant)
    - content: 메시지 내용
    - sql: 생성된 SQL (assistant 메시지)
    - rowCount: 쿼리 결과 건수 (assistant 메시지)
    - whereConditions: WHERE 조건 목록 (assistant 메시지) - Phase 1: 명시적 저장
    """
    if not conversation_history:
        return []

    sql_history = []
    for msg in conversation_history[-10:]:  # 최근 10개만
        entry: Dict[str, Any] = {
            "role": msg.role,
            "content": msg.content
        }

        # assistant 메시지에 SQL 정보가 있으면 포함
        if msg.role == "assistant" and msg.queryPlan:
            if msg.queryPlan.get("mode") == "text_to_sql" and msg.queryPlan.get("sql"):
                sql = msg.queryPlan.get("sql")
                entry["sql"] = sql

                # Phase 1: WHERE 조건을 명시적으로 추출하여 저장
                # 이를 통해 4단계+ 체이닝에서도 조건이 유실되지 않음
                if ENABLE_TEXT_TO_SQL:
                    where_conditions = extract_where_conditions(sql)
                    if where_conditions:
                        entry["whereConditions"] = where_conditions

        # 결과 건수 추출 (queryResult의 metadata에서)
        if msg.role == "assistant" and msg.queryResult:
            metadata = msg.queryResult.get("metadata", {})
            # totalRows 또는 rowsReturned 우선순위로 확인
            row_count = (
                msg.queryResult.get("totalCount") or
                metadata.get("totalRows") or
                metadata.get("rowsReturned")
            )
            if row_count is not None:
                entry["rowCount"] = row_count

        sql_history.append(entry)

    return sql_history


# ============================================
# 차트 렌더링 감지 및 구성 (TC-001)
# ============================================
# NOTE: CHART_KEYWORDS, TABLE_KEYWORDS, TIME_FIELD_KEYWORDS는
#       app.constants.render_keywords에서 import됨


def _detect_render_type_from_message(message: str) -> Optional[str]:
    """사용자 메시지에서 렌더링 타입 감지

    상수 파일(render_keywords.py)에서 키워드를 import하여 사용

    우선순위:
    1. 테이블 키워드 ("그래프 말고 표로" 같은 부정 표현 처리)
    2. 차트 키워드 (단독 키워드 "그래프", "차트" 포함)

    Args:
        message: 사용자 질문

    Returns:
        "chart" | "table" | None
    """
    msg = message.lower()

    # 1순위: 테이블 키워드 감지 (부정 표현 처리를 위해 먼저 체크)
    if any(kw in msg for kw in TABLE_KEYWORDS):
        return "table"

    # 2순위: 차트 키워드 감지 (단독 키워드 포함)
    if any(kw in msg for kw in CHART_KEYWORDS):
        return "chart"

    return None


def _detect_chart_type(data: List[Dict[str, Any]], columns: List[str], user_message: str = "") -> str:
    """데이터 구조를 분석하여 적절한 차트 타입 결정 (폴백 로직)

    LLM 기반 차트 타입 결정 실패 시 사용되는 규칙 기반 폴백 로직.
    사용자 메시지의 키워드와 데이터 구조를 분석하여 차트 타입 결정.

    Args:
        data: 쿼리 결과 데이터
        columns: 컬럼 목록
        user_message: 사용자 질문 (키워드 분석용)

    Returns:
        "bar" | "line" | "pie"
    """
    # render_keywords에서 import된 상수 사용
    from app.constants.render_keywords import CHART_TYPE_KEYWORDS, DATE_FIELDS

    if not data or not columns:
        return "bar"

    message_lower = user_message.lower()

    # 시계열 컬럼 감지 (DATE_FIELDS 상수 사용 - camelCase, snake_case 모두 지원)
    has_time_column = any(
        col.lower() in [f.lower() for f in DATE_FIELDS]
        or any(kw in col.lower() for kw in TIME_FIELD_KEYWORDS)
        for col in columns
    )

    # line 키워드 체크 (추이, 변화, 트렌드 등)
    line_keywords = CHART_TYPE_KEYWORDS.get("line", [])
    has_line_keyword = any(kw in message_lower for kw in line_keywords)

    # 시계열 + line 키워드 → line (데이터 행 수와 무관)
    if has_time_column and has_line_keyword:
        logger.info(f"[ChartType Fallback] time_column + line_keyword → line")
        return "line"

    # 시계열 + 2행 이상 → line (기존 임계값 완화: >2 → >=2)
    if has_time_column and len(data) >= 2:
        logger.info(f"[ChartType Fallback] time_column + data>=2 → line")
        return "line"

    # pie 키워드 → pie (10행 이하일 때만)
    pie_keywords = CHART_TYPE_KEYWORDS.get("pie", [])
    if any(kw in message_lower for kw in pie_keywords) and len(data) <= 10:
        logger.info(f"[ChartType Fallback] pie_keyword + data<=10 → pie")
        return "pie"

    # 카테고리가 적고 (5개 이하) 단일 값 컬럼이면 pie 차트
    # 단, line 키워드가 없는 경우에만
    if len(data) <= 5 and len(columns) == 2 and not has_line_keyword:
        logger.info(f"[ChartType Fallback] small_data + 2_cols + no_line_keyword → pie")
        return "pie"

    # 기본은 bar 차트
    logger.info(f"[ChartType Fallback] default → bar")
    return "bar"


def _identify_axis_keys(data: List[Dict[str, Any]], columns: List[str]) -> Tuple[str, str]:
    """X축과 Y축에 사용할 키 식별

    Args:
        data: 쿼리 결과 데이터
        columns: 컬럼 목록

    Returns:
        (x_key, y_key) 튜플
    """
    if not columns:
        return ("", "")

    if len(columns) == 1:
        return (columns[0], columns[0])

    # 숫자형 컬럼 찾기 (Y축 후보)
    numeric_cols = []
    category_cols = []

    if data:
        first_row = data[0]
        for col in columns:
            value = first_row.get(col)
            if isinstance(value, (int, float)):
                numeric_cols.append(col)
            else:
                category_cols.append(col)

    # X축: 카테고리/시간 컬럼, Y축: 숫자 컬럼
    x_key = category_cols[0] if category_cols else columns[0]
    y_key = numeric_cols[0] if numeric_cols else columns[-1]

    return (x_key, y_key)


def _detect_trend(values: List[float]) -> Optional[str]:
    """시계열 데이터의 추세 감지

    Args:
        values: Y축 값 리스트 (시간 순서대로)

    Returns:
        "증가" | "감소" | "유지" | None (데이터 부족시)
    """
    if len(values) < 3:
        return None

    # 전반부와 후반부의 평균 비교
    mid = len(values) // 2
    first_half = sum(values[:mid]) / mid
    second_half = sum(values[mid:]) / (len(values) - mid)

    if first_half == 0:
        return "증가" if second_half > 0 else "유지"

    diff_ratio = (second_half - first_half) / first_half

    if diff_ratio > 0.1:
        return "증가"
    elif diff_ratio < -0.1:
        return "감소"
    return "유지"


def _generate_insight(
    data: List[Dict[str, Any]],
    x_key: str,
    y_key: str,
    chart_type: str,
    template: Optional[str] = None
) -> Dict[str, Any]:
    """차트 데이터에 대한 인사이트 생성 (LLM 템플릿 우선, 규칙 기반 폴백)

    Args:
        data: 차트 데이터
        x_key: X축 필드 키
        y_key: Y축 필드 키
        chart_type: 차트 타입 (line, bar, pie)
        template: LLM이 생성한 인사이트 템플릿 (선택적)

    Returns:
        {
            "content": "인사이트 텍스트",
            "source": "llm" | "template" | "none"
        }
    """
    if not data:
        return {"content": None, "source": "none"}

    # 숫자 값 추출
    values = []
    for row in data:
        val = row.get(y_key, 0)
        if isinstance(val, (int, float)):
            values.append(float(val))
        elif val is not None:
            try:
                values.append(float(val))
            except (ValueError, TypeError):
                values.append(0)

    if not values:
        return {"content": None, "source": "none"}

    # 통계 계산
    count = len(data)
    total = sum(values)
    avg = total / count if count > 0 else 0
    max_val = max(values)
    min_val = min(values)

    # 최대/최소 카테고리 찾기
    max_idx = values.index(max_val)
    min_idx = values.index(min_val)
    max_category = str(data[max_idx].get(x_key, ""))
    min_category = str(data[min_idx].get(x_key, ""))

    # 추세 감지 (line 차트에서만)
    trend = _detect_trend(values) if chart_type == "line" else None

    # 필드 라벨 매핑 (snake_case → 한글)
    FIELD_LABELS = {
        "month": "월",
        "date": "일",
        "week": "주",
        "year": "연도",
        "day": "일",
        "merchant_id": "가맹점",
        "status": "상태",
        "method": "결제수단",
        "amount": "금액",
        "total_amount": "매출",
        "sum_amount": "총금액",
        "count": "건수",
        "payment_count": "결제건수",
        "refund_count": "환불건수",
        "avg_amount": "평균금액",
        "net_amount": "정산금액",
        "total": "합계",
        "avg": "평균",
    }

    # 필드 라벨 추출 (snake_case, camelCase 모두 지원)
    def get_field_label(key: str) -> str:
        key_lower = key.lower()
        if key_lower in FIELD_LABELS:
            return FIELD_LABELS[key_lower]
        # snake_case 처리
        parts = key.split('_')
        for part in parts:
            if part.lower() in FIELD_LABELS:
                return FIELD_LABELS[part.lower()]
        # 기본값: 그대로 반환
        return key.replace('_', ' ').title()

    group_by_label = get_field_label(x_key)
    metric_label = get_field_label(y_key)

    # 금액 포맷팅 함수
    def format_currency(val: float) -> str:
        if val >= 1000:
            return f"₩{int(val):,}"
        return f"{val:,.1f}"

    # 플레이스홀더 값 구성
    placeholders = {
        "{count}": f"{count:,}",
        "{total}": format_currency(total),
        "{avg}": format_currency(avg),
        "{max}": format_currency(max_val),
        "{min}": format_currency(min_val),
        "{maxCategory}": max_category,
        "{minCategory}": min_category,
        "{trend}": trend or "",
        "{groupBy}": group_by_label,
        "{metric}": metric_label,
    }

    # LLM 템플릿이 있으면 플레이스홀더 치환
    if template:
        content = template
        for placeholder, value in placeholders.items():
            content = content.replace(placeholder, value)

        # 미치환 플레이스홀더 제거 (중괄호로 시작하는 항목)
        import re
        content = re.sub(r'\{[^}]+\}', '', content)
        content = re.sub(r'\s+', ' ', content).strip()

        logger.info(f"[Insight] Generated from LLM template: {content[:100]}...")
        return {"content": content, "source": "llm"}

    # 폴백: 규칙 기반 템플릿
    if chart_type == "line":
        content = f"{group_by_label}별 {metric_label} 추이입니다. 총 {count:,}개 데이터의 합계는 {format_currency(total)}입니다."
        if trend:
            content += f" 전반적으로 {trend} 추세입니다."
    elif chart_type == "pie":
        content = f"{group_by_label}별 {metric_label} 분포입니다. {max_category}가 가장 큰 비중을 차지합니다."
    else:  # bar
        content = f"{group_by_label}별 {metric_label} 비교입니다. {max_category}가 {format_currency(max_val)}로 가장 높습니다."

    logger.info(f"[Insight] Generated from rule-based template: {content[:100]}...")
    return {"content": content, "source": "template"}


def _compose_chart_render_spec(
    result: Dict[str, Any],
    question: str,
    llm_chart_type: Optional[str] = None,
    insight_template: Optional[str] = None
) -> Dict[str, Any]:
    """차트 타입의 RenderSpec 구성

    Args:
        result: SQL 실행 결과
        question: 사용자 질문
        llm_chart_type: LLM이 추천한 차트 타입 (우선 사용)
        insight_template: LLM이 생성한 인사이트 템플릿 (선택적)

    Returns:
        차트 타입 RenderSpec (insight 필드 포함)
    """
    data = result.get("data", [])
    row_count = result.get("rowCount", 0)

    if not data:
        return {
            "type": "text",
            "title": "차트 생성 불가",
            "text": {
                "content": "조회 결과가 없어 차트를 생성할 수 없습니다.",
                "format": "plain"
            },
            "metadata": {
                "sql": result.get("sql"),
                "executionTimeMs": result.get("executionTimeMs"),
                "mode": "text_to_sql"
            }
        }

    columns = list(data[0].keys())

    # LLM 추천 차트 타입 우선 사용, 없으면 규칙 기반 폴백
    if llm_chart_type and llm_chart_type in ["line", "bar", "pie"]:
        chart_type = llm_chart_type
        logger.info(f"[ChartType] Using LLM recommendation: {chart_type}")
    else:
        # 폴백: 규칙 기반 로직 (개선 버전 - user_message 전달)
        chart_type = _detect_chart_type(data, columns, question)
        logger.info(f"[ChartType] Fallback to rule-based: {chart_type}")

    x_key, y_key = _identify_axis_keys(data, columns)

    # X축 라벨 생성
    x_label = x_key.replace("_", " ").title()
    y_label = y_key.replace("_", " ").title()

    # 차트 타입별 제목
    chart_type_names = {
        "bar": "막대 그래프",
        "line": "추이 그래프",
        "pie": "파이 차트"
    }
    title = f"{chart_type_names.get(chart_type, '차트')} ({row_count}건)"

    render_spec = {
        "type": "chart",
        "title": title,
        "chart": {
            "chartType": chart_type,
            "dataRef": "data.rows",
            "xAxis": {
                "dataKey": x_key,
                "label": x_label,
                "type": "category" if chart_type != "line" else "time"
            },
            "yAxis": {
                "dataKey": y_key,
                "label": y_label,
                "type": "number"
            },
            "series": [
                {
                    "dataKey": y_key,
                    "name": y_label,
                    "type": chart_type if chart_type in ["bar", "line"] else "bar"
                }
            ],
            "legend": True,
            "tooltip": True
        },
        "data": data,
        "metadata": {
            "sql": result.get("sql"),
            "executionTimeMs": result.get("executionTimeMs"),
            "mode": "text_to_sql",
            "chartType": chart_type
        }
    }

    # pie 차트의 경우 series 대신 별도 설정
    if chart_type == "pie":
        render_spec["chart"]["series"] = [
            {
                "dataKey": y_key,
                "name": y_label
            }
        ]

    # 인사이트 생성 및 추가
    insight = _generate_insight(
        data=data,
        x_key=x_key,
        y_key=y_key,
        chart_type=chart_type,
        template=insight_template
    )
    render_spec["chart"]["insight"] = insight

    return render_spec


def _escape_markdown_table_cell(value: str) -> str:
    """Markdown 테이블 셀의 특수문자 escape 처리

    테이블이 깨지지 않도록 파이프(|), 백틱(`) 등 처리
    """
    if value is None:
        return "-"
    s = str(value)
    # 파이프 문자는 테이블 구분자와 충돌하므로 escape
    s = s.replace("|", "\\|")
    # 줄바꿈은 공백으로 대체
    s = s.replace("\n", " ").replace("\r", "")
    return s


def _format_aggregation_as_markdown_table(
    row: Dict[str, Any],
    aggregation_context: Optional[Dict[str, Any]] = None
) -> str:
    """집계 결과를 Markdown 테이블로 변환

    Args:
        row: 집계 결과 단일 행 (예: {"total_amount": 14477000, "fee": 86862})
        aggregation_context: 집계 컨텍스트 (humanizedFilters 포함)

    Returns:
        Markdown 형식 문자열
    """
    # 컬럼명 → 한글 라벨 매핑
    COLUMN_LABELS = {
        # SQL 집계 함수 결과명 (PostgreSQL 기본 반환명)
        "sum": "합계",
        "count": "건수",
        "avg": "평균",
        "max": "최대값",
        "min": "최소값",
        # 별칭을 가진 집계 결과
        "original_amount": "원금액",
        "fee": "수수료",
        "amount_excluding_fee": "수수료 제외 금액",
        "total_amount": "총 금액",
        "total_fee": "총 수수료",
        "avg_amount": "평균 금액",
        "average_amount": "평균 금액",
        "max_amount": "최대 금액",
        "min_amount": "최소 금액",
        "sum_amount": "합계 금액",
        "payment_count": "결제 건수",
        "refund_count": "환불 건수",
        "net_amount": "정산 금액",
        "total_payment_amount": "총 결제 금액",
        "total_refund_amount": "총 환불 금액",
        # LLM이 자주 생성하는 별칭
        "completed_payment_count": "완료 결제 건수",
        "total_payments": "총 결제 건수",
        "avg_payment": "평균 결제 금액",
        "canceled_count": "취소 건수",
        "failed_count": "실패 건수",
        "total_sales": "총 매출",
        "total_transactions": "총 거래 건수",
        # 일반 컬럼명
        "amount": "금액",
        "merchant_id": "가맹점 ID",
        "status": "상태",
        "method": "결제수단",
    }

    # 금액 관련 키워드 (통화 포맷팅 적용)
    AMOUNT_KEYWORDS = ["amount", "fee", "total", "sum", "price", "balance", "net"]

    def format_value(key: str, value) -> str:
        """값을 포맷팅 (금액은 통화 형식, 건수는 "건" 접미사)"""
        from decimal import Decimal

        if value is None:
            return "-"

        # 숫자 타입 체크 (int, float, Decimal, 숫자 문자열)
        numeric_value = None
        if isinstance(value, (int, float, Decimal)):
            numeric_value = float(value)
        elif isinstance(value, str):
            try:
                numeric_value = float(value)
            except ValueError:
                pass

        if numeric_value is not None:
            int_val = int(numeric_value)
            # 금액 관련 필드면 통화 포맷
            if any(kw in key.lower() for kw in AMOUNT_KEYWORDS):
                return f"₩{int_val:,}"
            # count 필드면 "건" 접미사
            elif "count" in key.lower():
                return f"{int_val:,}건"
            else:
                return f"{int_val:,}"

        return _escape_markdown_table_cell(value)

    def get_label(key: str) -> str:
        """컬럼명을 한글 라벨로 변환"""
        if key in COLUMN_LABELS:
            return COLUMN_LABELS[key]
        # 스네이크 케이스를 공백으로 변환하고 Title Case 적용
        return key.replace("_", " ").title()

    # Markdown 테이블 생성
    lines = [
        "## 📊 집계 결과\n",
        "| 항목 | 값 |",
        "|------|------|"
    ]

    for key, value in row.items():
        label = get_label(key)
        formatted = format_value(key, value)
        # escape 처리된 라벨과 값 사용
        safe_label = _escape_markdown_table_cell(label)
        lines.append(f"| {safe_label} | {formatted} |")

    # 구분선
    lines.append("\n---\n")

    # 조회 조건 (humanized 사용)
    if aggregation_context:
        humanized_filters = aggregation_context.get("humanizedFilters", [])
        based_on_filters = aggregation_context.get("basedOnFilters", [])

        # humanizedFilters 우선, 없으면 basedOnFilters 사용
        filters_to_show = humanized_filters if humanized_filters else based_on_filters

        if filters_to_show:
            lines.append("**📌 조회 조건**")
            for filter_desc in filters_to_show:
                safe_filter = _escape_markdown_table_cell(filter_desc)
                lines.append(f"- {safe_filter}")
            lines.append("")

        # 기타 정보
        info_items = []
        source_count = aggregation_context.get("sourceRowCount")
        if source_count is not None:
            info_items.append(f"- 대상 데이터: {source_count:,}건")

        query_type = aggregation_context.get("queryType")
        if query_type:
            qtype_label = "새 쿼리 실행" if query_type == "NEW_QUERY" else "조건 추가"
            info_items.append(f"- 처리 방식: {qtype_label}")

        if info_items:
            lines.append("**📌 기타 정보**")
            lines.extend(info_items)

    return "\n".join(lines)


def compose_sql_render_spec(
    result: Dict[str, Any],
    question: str,
    llm_chart_type: Optional[str] = None,
    insight_template: Optional[str] = None
) -> Dict[str, Any]:
    """SQL 실행 결과를 RenderSpec으로 변환

    - 차트 요청: 차트 RenderSpec 반환 (TC-001)
    - 1000건 초과: 다운로드 RenderSpec (테이블 표시 안함)
    - 1000건 이하: 미리보기 10건 + 전체보기 모달

    Args:
        result: SQL 실행 결과
        question: 사용자 질문
        llm_chart_type: LLM이 추천한 차트 타입 (선택적)
        insight_template: LLM이 생성한 인사이트 템플릿 (선택적)
    """
    # TC-001: 차트 렌더링 타입 감지
    # LLM이 유효한 차트 타입을 추천했으면 차트로 렌더링
    render_type = _detect_render_type_from_message(question)

    # LLM 차트 타입이 유효하면(none이 아니면) 차트 요청으로 처리
    if llm_chart_type and llm_chart_type in ["line", "bar", "pie"]:
        logger.info(f"[compose_sql_render_spec] LLM chart type detected: {llm_chart_type}")
        return _compose_chart_render_spec(result, question, llm_chart_type, insight_template)

    # 메시지에서 차트 키워드 감지
    if render_type == "chart":
        return _compose_chart_render_spec(result, question, llm_chart_type, insight_template)

    data = result.get("data", [])
    row_count = result.get("rowCount", 0)
    total_count = result.get("totalCount") or row_count
    is_truncated = result.get("isTruncated", False)
    PREVIEW_LIMIT = 10  # 미리보기 행 수
    MAX_DISPLAY_ROWS = 1000  # 화면 표시 최대 건수

    # 1000건 초과: 다운로드 RenderSpec 반환
    if is_truncated:
        return {
            "type": "download",
            "title": "대용량 데이터 조회",
            "download": {
                "totalRows": total_count,
                "maxDisplayRows": MAX_DISPLAY_ROWS,
                "message": f"조회 결과가 {total_count:,}건으로 화면 표시 제한({MAX_DISPLAY_ROWS:,}건)을 초과합니다.",
                "sql": result.get("sql"),
                "formats": ["csv"]
            },
            "metadata": {
                "sql": result.get("sql"),
                "executionTimeMs": result.get("executionTimeMs"),
                "mode": "text_to_sql"
            }
        }

    if not data:
        return {
            "type": "text",
            "title": "조회 결과",
            "text": {
                "content": "조회 결과가 없습니다.",
                "format": "plain"
            },
            "metadata": {
                "sql": result.get("sql"),
                "executionTimeMs": result.get("executionTimeMs")
            }
        }

    # 집계 컨텍스트 추출
    is_aggregation = result.get("isAggregation", False)
    aggregation_context = result.get("aggregationContext")

    # 단일 행 + 집계 결과처럼 보이면 Markdown 테이블로 표시
    if row_count == 1 and len(data[0]) <= 5:
        row = data[0]
        # Markdown 테이블 + 조회 조건 생성
        content = _format_aggregation_as_markdown_table(row, aggregation_context)

        return {
            "type": "text",
            "title": "집계 결과",
            "text": {
                "content": content,
                "format": "markdown"
            },
            "metadata": {
                "sql": result.get("sql"),
                "executionTimeMs": result.get("executionTimeMs"),
                "mode": "text_to_sql",
                "isAggregation": is_aggregation,
                "aggregationContext": aggregation_context
            }
        }

    # 다중 행: 테이블로 표시 (미리보기 모드)
    if data:
        columns = list(data[0].keys())
        column_defs = []
        for col in columns:
            col_def = {
                "key": col,  # UI TableRenderer 호환
                "label": col.replace("_", " ").title(),  # TableRenderer expects 'label'
                "field": col,
                "headerName": col.replace("_", " ").title()
            }
            # 금액 필드 감지
            if any(kw in col.lower() for kw in ["amount", "fee", "net", "total", "price"]):
                col_def["type"] = "currency"
                col_def["currencyCode"] = "KRW"
            # 날짜 필드 감지
            elif any(kw in col.lower() for kw in ["date", "time", "at", "created", "updated"]):
                col_def["type"] = "datetime"
            column_defs.append(col_def)

        # 미리보기용 데이터 (최대 PREVIEW_LIMIT건)
        preview_data = data[:PREVIEW_LIMIT]
        has_more = row_count > PREVIEW_LIMIT

        # 타이틀: 미리보기인 경우 표시
        if has_more:
            title = f"조회 결과 ({row_count}건 중 {PREVIEW_LIMIT}건 미리보기)"
        else:
            title = f"조회 결과 ({row_count}건)"

        return {
            "type": "table",
            "title": title,
            # TC-004: 최상위 pagination 추가
            "pagination": {
                "totalRows": row_count,
                "totalPages": math.ceil(row_count / PREVIEW_LIMIT) if row_count > 0 else 1,
                "pageSize": PREVIEW_LIMIT,
                "hasMore": has_more
            },
            "table": {
                "columns": column_defs,
                "data": preview_data,  # 미리보기만 전송
                "dataRef": "data.rows",
                "actions": [
                    {"action": "fullscreen", "label": "전체보기"},
                    {"action": "export-csv", "label": "CSV 다운로드"}
                ],
                "pagination": {
                    "enabled": False,  # 미리보기에서는 페이지네이션 비활성화
                    "pageSize": PREVIEW_LIMIT,
                    "totalRows": row_count
                }
            },
            # 전체 데이터는 별도로 저장 (모달에서 사용)
            "fullData": data if has_more else None,
            "preview": {
                "enabled": has_more,
                "previewRows": PREVIEW_LIMIT,
                "totalRows": row_count,
                "message": f"전체 {row_count}건 중 {PREVIEW_LIMIT}건만 표시됩니다. 전체보기 버튼을 클릭하세요."
            },
            "metadata": {
                "sql": result.get("sql"),
                "executionTimeMs": result.get("executionTimeMs"),
                "mode": "text_to_sql"
            }
        }

    return {
        "type": "text",
        "title": "결과",
        "text": {
            "content": f"조회 완료: {row_count}건",
            "format": "plain"
        }
    }
