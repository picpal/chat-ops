"""
DownloadService: 대용량 쿼리 결과 다운로드

CSV, Excel 형식으로 쿼리 결과를 스트리밍 다운로드 지원
"""

import io
import csv
import logging
from typing import Generator, TYPE_CHECKING

if TYPE_CHECKING:
    from app.services.text_to_sql import TextToSqlService

logger = logging.getLogger(__name__)


def generate_csv(
    text_to_sql: "TextToSqlService",
    sql: str
) -> Generator[str, None, None]:
    """CSV 스트리밍 생성기

    대용량 데이터를 메모리 효율적으로 CSV로 변환합니다.
    배치 단위(1000건)로 데이터를 읽어 스트리밍합니다.

    Args:
        text_to_sql: TextToSqlService 인스턴스 (DB 연결용)
        sql: 실행할 SQL 쿼리 (LIMIT/OFFSET 제거된 상태)

    Yields:
        CSV 형식의 문자열 (헤더 + 데이터 행들)
    """
    import psycopg

    try:
        with text_to_sql._get_readonly_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(sql)

                # 헤더 출력
                columns = [desc[0] for desc in cur.description]
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(columns)
                yield output.getvalue()

                # 데이터 배치 처리 (1000건씩)
                batch_size = 1000
                row_count = 0
                while True:
                    rows = cur.fetchmany(batch_size)
                    if not rows:
                        break

                    output = io.StringIO()
                    writer = csv.writer(output)
                    for row in rows:
                        # datetime 변환
                        processed_row = []
                        for value in row.values() if hasattr(row, 'values') else row:
                            if hasattr(value, 'isoformat'):
                                processed_row.append(value.isoformat())
                            else:
                                processed_row.append(value)
                        writer.writerow(processed_row)
                        row_count += 1

                    yield output.getvalue()

                logger.info(f"CSV download completed: {row_count} rows")

    except psycopg.Error as e:
        logger.error(f"CSV download SQL execution failed: {e}")
        yield f"Error: {str(e)}"


def generate_excel(
    text_to_sql: "TextToSqlService",
    sql: str
) -> bytes:
    """Excel 파일 생성 (메모리 내)

    openpyxl을 사용하여 스타일이 적용된 Excel 파일을 생성합니다.
    헤더 행에 배경색과 볼드 폰트가 적용됩니다.

    Args:
        text_to_sql: TextToSqlService 인스턴스 (DB 연결용)
        sql: 실행할 SQL 쿼리 (LIMIT/OFFSET 제거된 상태)

    Returns:
        Excel 파일 바이트 데이터

    Raises:
        Exception: SQL 실행 실패 또는 Excel 생성 실패 시
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import psycopg
    import json

    wb = Workbook()
    ws = wb.active
    ws.title = "Query Result"

    try:
        with text_to_sql._get_readonly_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(sql)

                # 헤더 행 스타일
                columns = [desc[0] for desc in cur.description]
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

                for col_idx, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # 데이터 배치 처리 (1000건씩)
                batch_size = 1000
                row_num = 2
                while True:
                    rows = cur.fetchmany(batch_size)
                    if not rows:
                        break

                    for row in rows:
                        values = row.values() if hasattr(row, 'values') else row
                        for col_idx, value in enumerate(values, 1):
                            # datetime을 ISO format 문자열로 변환
                            if hasattr(value, 'isoformat'):
                                value = value.isoformat()
                            # dict/list (JSONB)를 JSON 문자열로 변환
                            elif isinstance(value, (dict, list)):
                                value = json.dumps(value, ensure_ascii=False, default=str)
                            ws.cell(row=row_num, column=col_idx, value=value)
                        row_num += 1

                logger.info(f"Excel generation completed: {row_num - 2} rows")

    except psycopg.Error as e:
        logger.error(f"Excel SQL execution failed: {e}")
        raise Exception(f"Excel generation failed: {str(e)}")

    # BytesIO로 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
